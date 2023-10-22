##########################################################
#                                                        #
#               develop by:- olakiya het                 #
#               email:- olakiyahet@gmail.com             #
#               github:- HETOLAKIYA007                   #            
##########################################################




import openpyxl

path = r"./Book1.xlsx"  # relative path
phoneNumbers = {}
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

for i in range(2, sheet_obj.max_row + 1):
    cell_obj2 = sheet_obj.cell(row=i, column=2)
    cell_obj3 = sheet_obj.cell(row=i, column=3)

    if cell_obj2.value in phoneNumbers:
        phoneNumbers[cell_obj2.value] += cell_obj3.value
    else:
        phoneNumbers[cell_obj2.value] = cell_obj3.value

    print(f"num:-{cell_obj2.value} min:-{cell_obj3.value}")

for num in phoneNumbers:
    total_minutes = phoneNumbers[num]
    print(f"{num}: {total_minutes}")

    if total_minutes < 60:
        print(f"{total_minutes} minutes")
    else:
        hours = total_minutes // 60
        minutes = total_minutes % 60
        print(f"{hours} hours and {minutes} minutes")

    

